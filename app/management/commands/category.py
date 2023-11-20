from django.core.management.base import BaseCommand
import openpyxl

from ...models import Category


class Command(BaseCommand):
    help = '1111111111111111111111111111111'

    def handle(self, *args, **options):
        data = openpyxl.load_workbook('KATEGORIYA.xlsx')
        sheet = data['Sheet1']
        for i in range(1, 680):
            icon = sheet.cell(row=i, column=1).value
            parent = sheet.cell(row=i, column=2).value
            if parent not in ('', None):
                paren_cat = Category.objects.create(name=parent, icon=icon)
            child = sheet.cell(row=i, column=3).value
            if child not in ('', None):
                Category.objects.create(name=child, parent=paren_cat, icon=icon)
        self.stdout.write(self.style.SUCCESS('MUVAFFAQIYATLI YAKUNLANDI'))

